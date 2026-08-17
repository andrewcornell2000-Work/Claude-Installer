"""Slop gate - lint a planned change before it is written.

    python lint_spec.py <workbook.xlsx> <spec.json> [spec.json ...]

Reads op-specs of the shape used by a spreadsheet applier (ops carrying `sheet`, `cell`/`range`,
`value`/`formula`/`text`) and refuses the ones that would write slop. Exits non-zero on any
REJECT so it can sit in front of the apply step.

Portable: copy into a project's _scripts/ and point it at that project's specs. The only
dependency is openpyxl, and only for the empty-row-reference check - without a workbook argument
every other rule still runs.
"""
import json, os, re, sys

sys.stdout.reconfigure(encoding="utf-8")

MAX_LABEL = 48

PROSE = re.compile(
    r"how to use|type only|the shaded|boxed cell|do not type|do not overwrite|owner input"
    r"|note:|memo:|story per|basis / source|leftover|not used|derived machinery"
    r"|for reference|\be\.g\.|\bi\.e\.", re.I)

# Constants a formula may carry without comment. Everything else is a candidate input.
BENIGN = {"0", "1", "-1", "12", "100", "2", "0.5"}
NUM_IN_FORMULA = re.compile(r"(?<![A-Z$:\d.])(\d+(?:\.\d+)?)(?![\d.]*[A-Z(])")
CELLREF = re.compile(r"(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_.]*))!\$?([A-Z]{1,3})\$?(\d+)")
UOM_OK = {"hours", "hours/day", "heads/day", "units", "cartons", "pallets", "$", "%", "fte",
          "index", "days", "weekends", "mixed", "steps", "year", "factor", "weeks", "all"}

rejects, warns = [], []

# Optional per-project conventions, read from .slopgate.json beside the specs. Absent means the
# convention-dependent rules stay quiet; the prose, label-length and empty-row rules always run.
#   {"uom_col_b_sheets": ["Variable Costs"], "input_sheets": ["Variable Assumptions", "CI"],
#    "max_label": 48}
CFG = {}


def texts_of(op):
    """Every human-visible string an op would write."""
    out = []
    for k in ("text", "value"):
        v = op.get(k)
        if isinstance(v, str):
            out.append((k, v))
    for v in op.get("values", []) or []:
        if isinstance(v, str):
            out.append(("values", v))
    return out


def lint(spec, wb):
    name = spec.get("_path", spec.get("spec", "?"))
    ops = spec.get("ops", [])
    # rows this spec writes a formula into, per sheet - used to spare a reference to a row the
    # same spec is about to populate
    being_written = {}
    for op in ops:
        addr = op.get("cell") or op.get("range") or ""
        mm = re.match(r"^\$?[A-Z]{1,3}\$?(\d+)", addr)
        if mm and op.get("sheet"):
            being_written.setdefault(op["sheet"], set()).add(int(mm.group(1)))
        if op.get("op") == "insert_rows":
            being_written.setdefault(op.get("sheet"), set()).update(op.get("rows", []))

    for i, op in enumerate(ops):
        where = f"{name} op#{i} {op.get('op')} {op.get('sheet','')}!{op.get('cell') or op.get('range') or op.get('row','')}"

        for kind, t in texts_of(op):
            s = t.strip()
            if not s or s.startswith("="):
                continue
            if PROSE.search(s):
                rejects.append((where, f"self-describing prose: {s[:70]!r}"))
            elif len(s) > MAX_LABEL:
                rejects.append((where, f"label {len(s)} chars, limit {MAX_LABEL}: {s[:70]!r}"))

        f = op.get("formula")
        if isinstance(f, str) and f.startswith("="):
            for lit in set(NUM_IN_FORMULA.findall(f)):
                if lit not in BENIGN:
                    warns.append((where, f"constant {lit} in a formula - should a human be able "
                                         f"to change it?"))
            if wb is not None:
                for q, b, col, row in CELLREF.findall(f):
                    sheet, row = (q or b), int(row)
                    if sheet not in wb.sheetnames:
                        continue
                    if row in being_written.get(sheet, ()):      # this spec creates it
                        continue
                    ws = wb[sheet]
                    if row > ws.max_row or all(
                            ws.cell(row, c).value in (None, "") for c in range(1, ws.max_column + 1)):
                        rejects.append((where, f"formula points at {sheet}!row {row}, which is "
                                               f"empty - a reference into blankness"))

        # These two depend on house conventions that vary by artifact, so they are opt-in. Run
        # them without configuring and every input cell and every non-UoM column B reports as a
        # defect, which buries the rules that matter. Declare the conventions in .slopgate.json.
        if CFG.get("input_sheets") and op.get("op") == "set" \
                and isinstance(op.get("value"), (int, float)) \
                and not isinstance(op.get("value"), bool) \
                and op.get("sheet") not in CFG["input_sheets"]:
            warns.append((where, f"typed number {op['value']} on a non-input sheet - should it be "
                                 f"an input cell a human can change?"))

        if op.get("op") == "insert_cols":
            warns.append((where, "a column insert cuts EVERY row of the sheet - list what else "
                                 "occupies those columns first"))

        if op.get("sheet") in CFG.get("uom_col_b_sheets", []) and op.get("op") == "set" \
                and str(op.get("cell", "")).startswith("B") \
                and isinstance(op.get("value"), str) \
                and op["value"].strip().lower() not in UOM_OK:
            warns.append((where, f"unit of measure {op['value']!r} is not one of the house set"))


def main():
    global MAX_LABEL
    args = sys.argv[1:]
    wb, specs = None, []
    for cand in args:
        cfg = os.path.join(os.path.dirname(os.path.abspath(cand)), ".slopgate.json")
        if os.path.exists(cfg):
            CFG.update(json.load(open(cfg, encoding="utf-8")))
            MAX_LABEL = CFG.get("max_label", MAX_LABEL)
            print(f"  conventions from {cfg}")
            break
    if args and args[0].lower().endswith((".xlsx", ".xlsm")):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(args[0], data_only=True)
        except Exception as e:                      # the other rules still apply
            print(f"  (no workbook loaded: {e}; empty-row check skipped)")
        args = args[1:]
    for p in args:
        s = json.load(open(p, encoding="utf-8"))
        s["_path"] = os.path.basename(p)
        specs.append(s)
    for s in specs:
        lint(s, wb)

    print("=" * 92)
    print(f"SLOP GATE - {len(specs)} spec(s), "
          f"{sum(len(s.get('ops', [])) for s in specs)} ops")
    print("=" * 92)
    for w, m in rejects:
        print(f"  REJECT  {w}\n          {m}")
    for w, m in warns:
        print(f"  warn    {w}\n          {m}")
    print(f"\n{len(rejects)} rejects, {len(warns)} warnings")
    if rejects:
        print("*** slop gate FAILED - fix the design, do not caption it ***")
    return 1 if rejects else 0


if __name__ == "__main__":
    sys.exit(main())
